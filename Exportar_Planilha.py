import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.drawing.image import Image
from datetime import datetime
from pathlib import Path
import xlwings as xw
import subprocess
import os

def copiar_formatacao(sheet, linha_origem, linha_destino):
    for col in range(3, 9):  # Limita às colunas de 4 a 8
        celula_origem = sheet.cell(row=linha_origem, column=col)
        celula_destino = sheet.cell(row=linha_destino, column=col)

        # Copia diretamente a formatação da célula de cima
        celula_destino._style = celula_origem._style

def exportar_para_excel(dados_extraidos):
    """Salva os dados extraídos em uma planilha Excel com data e hora no nome do arquivo."""
    '''
    print(f"\n{'='*80}")
    print(f"📊 INICIANDO EXPORTAÇÃO PARA EXCEL")
    print(f"{'='*80}")
    '''
    if not dados_extraidos:
        print("❌ ERRO: Nenhum dado para exportar.")
        return

    print(f"📈 Total de registros para exportar: {len(dados_extraidos)}")
    
    '''
    # DEBUG: Verificar tipos de dados
    print(f"\n🔍 DEBUG - Verificando tipos de dados:")
    for i, dados in enumerate(dados_extraidos):
        print(f"\n   📋 Registro {i+1} - UC: {dados.get('uc', 'N/A')}")
        
        # Verificar campos que podem causar erro
        campos_criticos = ['consumo', 'saldo', 'valor_economia', 'desconto_fatura', 'valor_consorcio']
        for campo in campos_criticos:
            if campo in dados:
                valor = dados[campo]
                tipo = type(valor).__name__
                print(f"      {campo}: {tipo} = {valor}")
    '''
    try:
        print(f"\n📝 Criando DataFrame...")
        # Cria um DataFrame com os dados extraídos
        df = pd.DataFrame(dados_extraidos)
        #print(f"✅ DataFrame criado com sucesso - Shape: {df.shape}")
    except Exception as e:
        print(f"❌ ERRO ao criar DataFrame: {e}")
        print(f"   Tipo do erro: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return

    # Define a pasta de destino
    pasta_pdfs = Path.home() / "Dropbox" / "AUPUS SMART" / "01. Club AUPUS" / "01. Usineiros" / "01. AUPUS ENERGIA" / "01. FATURAS" / "2025" / "2025.04"

    # Garante que a pasta exista
    try:
        os.makedirs(pasta_pdfs, exist_ok=True)
    except Exception as e:
        print(f"❌ ERRO ao verificar pasta: {e}")
        return

    # Gera um nome de arquivo com data e hora
    data_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nome_arquivo = f"output_{data_hora}.xlsx"
    caminho_saida = os.path.join(pasta_pdfs, nome_arquivo)

    try:
        # Salva a planilha
        df.to_excel(caminho_saida, index=False, sheet_name="Dados")
    except Exception as e:
        print(f"❌ ERRO ao salvar Excel: {e}")
        print(f"   Tipo do erro: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return

    # Agora, vamos manipular a planilha existente
    caminho_planilha = Path.home() / "Dropbox" / "AUPUS SMART" / "01. Club AUPUS" / "_Controles" / "06. Controles" / "AUPUS ENERGIA.xlsm"
    
    if not os.path.exists(caminho_planilha):
        print(f"❌ ERRO: Planilha de controle não encontrada: {caminho_planilha}")
        return
    
    # Itera sobre cada dicionário na lista de dados extraídos
    for i, dados in enumerate(dados_extraidos):
        print(f"\n🔄 Processando registro {i+1}/{len(dados_extraidos)}...")
        try:
            atualizar_planilha_controle(caminho_planilha, dados)
            print(f"✅ Registro {i+1} processado com sucesso")
        except Exception as e:
            print(f"❌ ERRO ao processar registro {i+1}: {e}")
            print(f"   UC: {dados.get('uc', 'N/A')}")
            print(f"   Tipo do erro: {type(e).__name__}")
            import traceback
            traceback.print_exc()
            
            # Continuar com próximo registro ao invés de parar
            continue
    
    print(f"\n✅ Exportação finalizada!")

def atualizar_planilha_controle(caminho_planilha, dados):
    """Atualiza a planilha de controle com os dados extraídos."""
    
    print(f"   📂 Abrindo planilha de controle...")

    try:
        # Abre a planilha existente
        workbook = openpyxl.load_workbook(caminho_planilha, keep_vba=True)
        sheet_dados = workbook["DADOS"]
        sheet_demo = workbook["DEMONSTRATIVO"]
        sheet_grafico = workbook["GRAFICO"]

    except Exception as e:
        print(f"   ❌ ERRO ao abrir planilha: {e}")
        raise e

    # Obtém a UC dos dados extraídos
    uc = dados.get("uc")
    if not uc:
        print("   ❌ ERRO: UC não encontrada nos dados.")
        workbook.close()
        return

    print(f"   🔍 Procurando UC: {uc}")

    linha_uc = None  # Inicializa como None para indicar que a UC ainda não foi encontrada
    try:
        for row in sheet_dados.iter_rows(min_row=2, max_row=sheet_dados.max_row, min_col=2, max_col=2):
            cell_value = str(row[0].value).strip() if row[0].value else ""
            if cell_value == str(uc):
                linha_uc = row[0].row
                break
    except Exception as e:
        print(f"   ❌ ERRO ao procurar UC: {e}")
        workbook.close()
        raise e
    
    if linha_uc is None:
        print(f"   ⚠️ UC {uc} não encontrada na planilha - pulando...")
        workbook.close()
        return

    try:
        # Define a linha inicial como 2 linhas acima da linha onde a UC foi encontrada
        linha_inicial = linha_uc - 2
        linha_teste = linha_inicial + 1

        # Se já está vazio, é ali mesmo que escreve
        if sheet_dados.cell(row=linha_teste, column=3).value is None:
            ultima_linha = linha_teste
        else:
            # Se não, procura a próxima linha vazia
            ultima_linha = linha_teste
            while sheet_dados.cell(row=ultima_linha, column=3).value is not None:
                ultima_linha += 1

        # Verifica se precisa adicionar nova linha
        if ultima_linha > linha_inicial + 9:
            sheet_dados.insert_rows(ultima_linha)

        # Linha para escrever
        linha_escrever = ultima_linha

        # Formata a data para "mes_escrito_abreviado/ano_abreviado"
        periodo_faturamento = dados.get("data_leitura")
        data_formatada = None
        data = None
        
        if periodo_faturamento:
            try:
                data = datetime.strptime(periodo_faturamento, "%d/%m/%y")
                mes_ano_abreviado = data.strftime("%b/%y").lower()
                sheet_dados.cell(row=linha_escrever, column=3, value=mes_ano_abreviado)
                data_formatada = datetime.strptime(periodo_faturamento, "%d/%m/%y").strftime("%d/%m/%Y")
                sheet_dados.cell(row=linha_escrever, column=4, value=data_formatada)
                print(f"   📅 Data processada: {periodo_faturamento} → {mes_ano_abreviado}")
            except ValueError as e:
                print(f"   ⚠️ Data inválida: {periodo_faturamento} - {e}")

        
        # DEBUG: Verificar tipos antes de escrever
        consumo = dados.get("consumo")
        saldo = dados.get("saldo")
        excedente_recebido = dados.get("excedente_recebido")
        valor_economia = dados.get("valor_economia")
        
        print(f"      consumo: {type(consumo).__name__} = {consumo}")
        print(f"      saldo: {type(saldo).__name__} = {saldo}")
        print(f"      excedente_recebido: {type(excedente_recebido).__name__} = {excedente_recebido}")
        print(f"      valor_economia: {type(valor_economia).__name__} = {valor_economia}")

        # Escreve os demais valores
        sheet_dados.cell(row=linha_escrever, column=5, value=consumo)
        sheet_dados.cell(row=linha_escrever, column=6, value=saldo)
        sheet_dados.cell(row=linha_escrever, column=7, value=excedente_recebido)
        sheet_dados.cell(row=linha_escrever, column=8, value=valor_economia)

        print(f"   ✅ Dados básicos escritos")

        endereco_cliente = sheet_dados.cell(row=linha_inicial+9, column=2).value
        dados["endereco_cliente"] = endereco_cliente
        cpf_cnpj_cliente = sheet_dados.cell(row=linha_inicial+1, column=2).value
        dados["cpf_cnpj_cliente"] = cpf_cnpj_cliente

        copiar_formatacao(sheet_dados, linha_escrever - 1, linha_escrever)
        
        # Calcula as médias na coluna 8
        valores_coluna_8 = [
            sheet_dados.cell(row=row, column=8).value for row in range(linha_inicial + 1, linha_escrever + 1)
            if sheet_dados.cell(row=row, column=8).value is not None
        ]

        # CORREÇÃO: Converter todos os valores para float antes de somar
        try:
            valores_float = []
            for val in valores_coluna_8:
                if val is not None:
                    if hasattr(val, '__float__'):  # Decimal, int, float
                        valores_float.append(float(val))
                    else:
                        valores_float.append(float(val))
            
            # Média dos últimos 12 valores (ou menos)
            soma_12 = sum(valores_float[-12:]) if valores_float else 0
            soma_12 = round(soma_12, 2)
            sheet_dados.cell(row=linha_inicial + 4, column=2, value=soma_12)

            # Média de todos os valores
            soma_total = sum(valores_float) if valores_float else 0
            soma_total = round(soma_total, 2)
            sheet_dados.cell(row=linha_inicial + 5, column=2, value=soma_total)

            
        except Exception as calc_err:
            print(f"      ❌ ERRO ao calcular médias: {calc_err}")
            # Valores padrão em caso de erro
            soma_12 = 0
            soma_total = 0
            sheet_dados.cell(row=linha_inicial + 4, column=2, value=soma_12)
            sheet_dados.cell(row=linha_inicial + 5, column=2, value=soma_total)

        # Escreve o saldo
        sheet_dados.cell(row=linha_inicial + 8, column=2, value=saldo)
        
        # Preenchimento da aba DEMONSTRATIVO
        sheet_demo.cell(row=2, column=2, value=dados.get("nome"))
        sheet_demo.cell(row=4, column=2, value=dados.get("cpf_cnpj_cliente"))
        sheet_demo.cell(row=6, column=2, value=dados.get("endereco_cliente"))
        sheet_demo.cell(row=10, column=2, value=dados.get("tipo_fornecimento"))
        sheet_demo.cell(row=12, column=2, value=dados.get("uc"))
        
        if data:
            mes_unitario = data.strftime("%m")
            sheet_demo.cell(row=14, column=2, value=mes_unitario)
            
        data_atual = datetime.now().strftime("%d/%m/%Y")
        sheet_demo.cell(row=16, column=2, value=data_atual)
        sheet_demo.cell(row=19, column=2, value=soma_12)
        sheet_demo.cell(row=22, column=2, value=soma_total)

        sheet_demo.cell(row=16, column=4, value=dados.get("aliquota_icms"))
        sheet_demo.cell(row=16, column=6, value=dados.get("aliquota_pis"))
        sheet_demo.cell(row=16, column=8, value=dados.get("aliquota_cofins"))

        sheet_demo.cell(row=19, column=4, value=dados.get("rs_adc_bandeira_amarela"))
        sheet_demo.cell(row=19, column=6, value=dados.get("rs_adc_bandeira_vermelha"))
        if data_formatada:
            sheet_demo.cell(row=19, column=8, value=data_formatada)

        sheet_demo.cell(row=22, column=4, value=dados.get("excedente_recebido"))
        sheet_demo.cell(row=22, column=6, value=dados.get("saldo"))
        sheet_demo.cell(row=22, column=8, value=dados.get("valor_concessionaria"))

        sheet_demo.cell(row=25, column=4, value=dados.get("consumo_comp"))
        sheet_demo.cell(row=25, column=6, value=dados.get("valor_comp"))
        sheet_demo.cell(row=25, column=8, value=dados.get("valor_band_comp"))

        sheet_demo.cell(row=28, column=4, value=dados.get("consumo_n_comp"))
        sheet_demo.cell(row=28, column=6, value=dados.get("valor_com_desconto"))
        sheet_demo.cell(row=28, column=8, value=dados.get("valor_bandeira_com_desconto"))

        sheet_demo.cell(row=31, column=4, value=dados.get("consumo"))
        sheet_demo.cell(row=31, column=6, value=dados.get("valor_aupus"))
        sheet_demo.cell(row=31, column=8, value=dados.get("valor_consorcio"))

        sheet_demo.cell(row=76, column=2, value=dados.get("desconto_fatura", 0))
        
        # Limpeza da aba GRAFICO
        for row in range(2, 14):
            for col in range(1, 3):
                sheet_grafico.cell(row=row, column=col).value = None

        # Limpar as linhas 16A até 27B
        for row in range(16, 28):
            for col in range(1, 4):
                sheet_grafico.cell(row=row, column=col).value = None

        linha_inicial = linha_inicial +1
        qtd_linhas = ultima_linha - linha_inicial + 1

        if qtd_linhas > 12:
            linha_inicial = ultima_linha - 11  # Mantém apenas os últimos 12 meses
            qtd_linhas = 12

        for i in range(qtd_linhas):
            valor = sheet_dados.cell(row=linha_inicial + i, column=3).value
            sheet_grafico.cell(row=2 + i, column=1).value = valor  # Linha 2 a 13
            sheet_grafico.cell(row=16 + i, column=1).value = valor  # Linha 16 a 27

        for i in range(qtd_linhas):
            valor = sheet_dados.cell(row=linha_inicial + i, column=5).value
            sheet_grafico.cell(row=16 + i, column=2).value = valor

        for i in range(qtd_linhas):
            valor = sheet_dados.cell(row=linha_inicial + i, column=6).value
            sheet_grafico.cell(row=16 + i, column=3).value = valor

        for i in range(qtd_linhas):
            valor = sheet_dados.cell(row=linha_inicial + i, column=8).value
            sheet_grafico.cell(row=2 + i, column=2).value = valor

        sheet_mes = workbook["SETEMBRO"]

        ultima_linha = 3
        while sheet_mes.cell(row=ultima_linha, column=2).value is not None:
            ultima_linha += 1

        linha_escrever = None  # Inicializa como None para indicar que a UC ainda não foi encontrada
        for row in sheet_mes.iter_rows(min_row=4, max_row=sheet_mes.max_row, min_col=3, max_col=3):
            cell_value = str(row[0].value).strip() if row[0].value else ""
            if cell_value == str(uc):
                linha_escrever = row[0].row
                break

        if linha_escrever is None:
            linha_escrever = ultima_linha

        print(f"   🧮 Calculando rateio...")
        
        # DEBUG: Verificar tipos antes do cálculo crítico
        consumo = dados.get("consumo")
        saldo = dados.get("saldo")

        try:
            # CORREÇÃO: Converter para float antes do cálculo
            consumo_float = float(consumo) if consumo is not None else 0.0
            saldo_float = float(saldo) if saldo is not None else 0.0
            
            # PONTO CRÍTICO: Esta linha pode estar causando o erro
            rateio = (consumo_float * 1.1) - saldo_float
            rateio = round(rateio, 2)
            if rateio < 0:
                rateio = 0
        except Exception as calc_err:
            print(f"      ❌ ERRO NO CÁLCULO DO RATEIO: {calc_err}")
            print(f"      Tipo do erro: {type(calc_err).__name__}")
            rateio = 0  # Valor padrão em caso de erro

        sheet_mes.cell(row=linha_escrever, column=1, value=dados.get("id_planilha"))
        nome = dados.get("nome")
        sheet_mes.cell(row=linha_escrever, column=2, value=nome)
        uc=dados.get("uc")
        sheet_mes.cell(row=linha_escrever, column=3, value=uc)
        sheet_mes.cell(row=linha_escrever, column=4, value=dados.get("desconto_fatura"))
        sheet_mes.cell(row=linha_escrever, column=5, value=dados.get("consumo_comp"))
        sheet_mes.cell(row=linha_escrever, column=6, value=dados.get("consumo_n_comp"))
        sheet_mes.cell(row=linha_escrever, column=7, value=dados.get("consumo"))
        sheet_mes.cell(row=linha_escrever, column=8, value=dados.get("valor_consorcio"))
        sheet_mes.cell(row=linha_escrever, column=9, value=dados.get("saldo"))
        sheet_mes.cell(row=linha_escrever, column=10, value=rateio)
        sheet_mes.cell(row=linha_escrever, column=11, value=dados.get("cpf_cnpj_cliente"))
        sheet_mes.cell(row=linha_escrever, column=12, value=dados.get("vencimento_consorcio"))
        
        img = Image(Path.home() / "Dropbox" / "AUPUS SMART" / "01. Club AUPUS" / "01. Usineiros" / "01. AUPUS ENERGIA" / "_Controles" / "NET.png")
        sheet_demo.add_image(img, 'D2')

        workbook.save(caminho_planilha)
        # Fechar corretamente para evitar corrompimento
        workbook.close()
        del workbook

        app = xw.App(visible=False)  # Mantém o Excel em segundo plano
        wb = xw.Book(caminho_planilha)  # Abre o arquivo salvo

        # Executa a macro
        wb.macro("GerarPDF")()

        # Salva e fecha a planilha
        wb.save()
        wb.close()
        app.quit()

    except Exception as e:
        print(f"   ❌ ERRO durante processamento: {e}")
        print(f"   Tipo do erro: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        
        # Tentar fechar a planilha em caso de erro
        try:
            workbook.close()
        except:
            pass
        raise e